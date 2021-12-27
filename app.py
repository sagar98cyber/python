import openpyxl as xl
from openpyxl import chart
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

#To get a single value from some cell
#print(sheet.cell(2,3).value)

#To find the maximum enteries in a row
#print(sheet.max_row)

#To find the maximum enteries in a column
#print(sheet.max_column)

for row in range(2,sheet.max_row + 1):
    #accessing the single cell
    cell_value = sheet.cell(row,3).value
    corrected_price = cell_value * 0.9
    #The next cell where corrected value is to be entered
    corrected_cell_value = sheet.cell(row,4)
    #Assigning the corrected value to the correct value cell
    corrected_cell_value.value = corrected_price

values = Reference(sheet,min_row = 2, max_row = sheet.max_row,min_col = 4, max_col = 4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
#saving all the changes in a new Excel File called 'transactions2.xlsx'
wb.save('transactions2.xlsx')